# 该工具用于采集图片和文字识别获取 - 物产志-提瓦特物产/战利品
# 设为通用
# 01. 【标准部件】判断是否前置窗口为原神get_window_handle()  get_width()
# 02. 【标准部件】利用pyautogui对当前窗口（全屏）进行截取，获取窗口化的窗口的实际坐标进行计算 软件部分截图：py_screenshot(screenshot_png,pic_save=True) 全屏截图：py_all_screenshot(screenshot_png,pic_save=True)
# 03. 【标准部件】对截图的指定部分进行截取crop_and_save_image(screenshot, x, y, width, height, save_path)
# 04. 【标准部件】对屏幕指定位置进行单击和中键移动py_click(x,y,width,height) py_scroll(rows)
# 05. 【标准部件】图片与图片进行比对，确定被比对的图在截图中的位置 find_image_coordinates(image_path, template_path)
# 06. 【标准部件】文字识别 text_ocr(pic_ocr)
# 07. 测试：对单个文字进行获取
# 07. 实操：物产志中循环文字提取并输出图片

"""
BUG: 1.当某个素材缺失后，无法进行有效文字识别
"""

import os
import time
import shutil
import keyboard
from PIL import Image

def on_key_release(event):
    if event.name == 'f3':
        # 按下 F3 键后终止程序
        print("按下 F3 键，程序退出。")
        os._exit(0)

# 注册按键释放事件的回调函数
keyboard.on_release(on_key_release)

'''
图片大小为114*114（与背包相比少10px，124x124）
第一个位置
左上 140,125

第二个
左上 269,125    差异129 横向


下方
左上 140,286  差异161

文字位置
左上    1322,666    高228
左下    1322,894                【下方文字位置为900高】
'''
#'''
print('执行01步骤***************************************')
print('########### 01. 【标准部件】判断是否前置窗口为原神')
from utils_find_process import get_window_handle
from utils_find_process import get_window_handle_name
from utils_find_process import get_foreground_window_info
from utils_process_x_y import get_width
from utils_process_x_y import get_process_dimensions

# 其他语言的标题不是原神！【【【【【【【【【【【【备忘问题】】】】】】】】】】】】
if get_window_handle('原神'):   # 将标题名为原神的进程前置
    # print('你都没开原神！还是改语言了？')
    lag = 'chinese'
else:
    get_window_handle_name('YuanShen.exe')
    time.sleep(0.5)
    handle, pid, name, title = get_foreground_window_info()
    if title == 'Genshin Impact':
        # print('not open Genshin Impact?')
        lag = 'english'
    else:
        lag = 'other'
    # os._exit(1)     # 没做其他语言判断
print(lag)
if lag == 'chinese':
    real_width , real_height ,scaling ,borderless = get_width('原神')
elif lag == 'english':
    real_width , real_height ,scaling ,borderless = get_process_dimensions("YuanShen.exe")
else:
    real_width , real_height ,scaling ,borderless = get_process_dimensions("YuanShen.exe")
time.sleep(0.5)
if not real_width == 1920 or not real_height == 1080:
    print('窗口不是1920X1080')
    os._exit(1)
#'''

print('执行02步骤***************************************')
print('########### 02. 【标准部件】利用pyautogui对当前窗口进行截取，获取窗口化的窗口的实际坐标进行计算')
# 问题：窗口必须在主界面，否则无法截取，显示为全黑
import pygetwindow as gw
import pyautogui
# 获取当前活动窗口
active_window = gw.getActiveWindow()            # <Win32Window left="0", top="0", width="1920", height="1080", title="原神">    # 偏移和尺寸
# 获取窗口位置和大小
window_left = active_window.left        # 相对于屏幕左上角的水平坐标
window_top = active_window.top          # 相对于屏幕左上角的垂直坐标
window_width = active_window.width      # 当前窗口的宽度
window_height = active_window.height    # 当前窗口的高度
# 进行截图
screenshot = pyautogui.screenshot(region=(window_left, window_top, window_width, window_height))        # 元组，window_left和window_top设为0则为全屏截图
# 当前窗口信息 <Win32Window left="1912", top="-8", width="1936", height="1056", title="dev2.py - 100 - Visual Studio Code [管理员]">
# 多屏下，为1屏时，x正确。在其他屏时，相对1屏的位置（从1屏开始算轴，不论1屏的位置）
print('当前窗口信息',active_window)
print('相对于屏幕左上角的水平坐标',window_left)
print('相对于屏幕左上角的垂直坐标',window_top)
print('当前窗口的宽度',window_width)
print('当前窗口的高度',window_height)

def py_screenshot(screenshot_png,pic_save=True):  # 截图：'test-screenshot.png'，仅限软件截图而非全屏截图,做坐标判断时，需要加入全屏偏移
    screenshot_pic = pyautogui.screenshot(region=(window_left, window_top, window_width, window_height))
    if pic_save == True:
        screenshot_pic.save(screenshot_png)     # 保存到图片，find_image_coordinates方法需要使用图片路径，否则需要加上判断
    return screenshot_pic       # 注意！输出的是图片的值，而不是图片的路径
# py_screenshot('test-screenshot.png')

def py_all_screenshot(screenshot_png,pic_save=True):  # 截图：'test-screenshot.png' ,全屏而不是单个软件部分，用于单击坐标获取
    screenshot_pic = pyautogui.screenshot()
    if pic_save == True:
        screenshot_pic.save(screenshot_png)     # 保存到图片
    return screenshot_pic       # 注意！输出的是图片的值，而不是图片的路径
# py_screenshot('test-screenshot.png')

print('执行03步骤***************************************')
print('########### 03. 【标准部件】对截图的指定部分进行截取')
def crop_and_save_image(screenshot, x, y, width, height, save_path):        # 从前端获得A. screenshot截图，B C.定义x、y坐标，D E.宽高矩形，F. 保存位置
    # 判断输入类型是图像路径还是图像对象
    if isinstance(screenshot, str):  # 如果是字符串，则加载图像
        screenshot = Image.open(screenshot)
    # 根据给定的坐标和尺寸进行截取
    cropped_image = screenshot.crop((x, y, x + width, y + height))
    # 保存截取的图片
    cropped_image.save(save_path)
# crop_and_save_image(screenshot, 140, 125, 114, 114, 'test-cropped_image.png')

print('执行04步骤***************************************')
print('########### 04. 【标准部件】对屏幕指定位置进行单击和中键移动')
# 单击操作
def py_click(x,y,width,height):     # x坐标，y坐标=左上角顶点坐标，width、height=宽高中心点增值，也可以设为0
    click_loc = (x + width + window_left,y + height + window_top)    # 左上角顶点+中心点增值 + 窗口与屏幕的差距
    # print('单击位置',click_loc)
    pyautogui.click(click_loc) 
    return click_loc
# py_click(140, 125, 114/2, 114/2)

# 滚动操作
def py_scroll(rows):
    for i in range(0,rows):                    # 右侧为绕过的行数（含本行）,物产志提瓦特物产首页5行-堇瓜
        for i in range(0,9):                # 素材的一行，物产志和素材一致,9次滚动
            pyautogui.scroll(-1,x=0,y=0)    # 只定义正值（向上）和负值（向下），x水平正值向右，负值向左。y正值向上，负值向下
    pyautogui.scroll(1,x=0,y=0)             # 返回向上一次
    time.sleep(0.5)
# py_scroll(5)    

print('执行05步骤***************************************')
print('########### 05. 【标准部件】图片与图片进行比对，确定被比对的图在截图中的位置')
import cv2
import numpy as np
def find_image_coordinates(image_path, template_path):      # image_path = 屏幕截图 template_path = 对比图,添加了，输入内容判断是图片还是图片的路径
    # 判断image_path和template_path是否为字符串路径
    if isinstance(image_path, str):
        # 读取图像文件
        image = cv2.imread(image_path)
    else:
        # image_path为图像变量
        image = np.array(image_path)
        image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

    if isinstance(template_path, str):
        # 读取模板文件
        template = cv2.imread(template_path)
    else:
        # template_path为图像变量
        template = np.array(template_path)
        template = cv2.cvtColor(template, cv2.COLOR_RGB2BGR)
    # 使用模板匹配算法查找模板在图像中的位置
    result = cv2.matchTemplate(image, template, cv2.TM_CCOEFF_NORMED)
    _, max_val, _, max_loc = cv2.minMaxLoc(result)      # max_val是相似度，max_loc为坐标（元组）
    # 获取模板的宽度和高度
    template_width = template.shape[1]
    template_height = template.shape[0]

    # 计算模板的左上角和右下角坐标
    top_left = max_loc      # 左上角坐标
    bottom_right = (top_left[0] + template_width, top_left[1] + template_height)    # 右下角坐标，依赖比对图的大小，如果有差异、拉伸则错误，114或125
    print('比对相似度为：',max_val, '坐标为：',top_left, bottom_right)
    return max_val,top_left, bottom_right

print('执行06步骤***************************************')
print('########### 06. 【标准部件】文字识别')
import cnocr
def text_ocr(pic_ocr,single=False):
    current_dir = os.getcwd()   # 当前目录，要把模型放在当前目录，model目录
    # cnocr
    #print('载入OCR模型')
    if lag == 'english':
        det_model_name = "en_PP-OCRv3_det"
    elif lag == 'chinese':
        det_model_name = "ch_PP-OCRv3_det"
    else:
        det_model_name = "en_PP-OCRv3_det"              # 【【【【【注意！其他语言没有指定OCR模型！可能全错】】】】】
    det_root = os.path.join(current_dir, "model/cnocr")
    # cnstd
    #print('载入文字识别模型')
    rec_model_name = "densenet_lite_114-fc"
    rec_root = os.path.join(current_dir, "model/cnstd")
    #print('文字识别启用')
    # 使用 OCR 模型进行文本识别，仅数字后面加上, cand_alphabet='0123456789'
    ocr = cnocr.CnOcr(det_model_name=det_model_name, rec_model_name=rec_model_name, det_root=det_root, rec_root=rec_root)
    if single == True:
        # 获取识别结果中的文本值 # 单行内容  
        text = ocr.ocr_for_single_line(pic_ocr)
    else:
        text = ocr.ocr(pic_ocr)
    return text

print('该部分判断头部分以便确认属于哪个类目的内容')
if lag == 'chinese':
    screenshot_pic = py_screenshot('./dev2/A-001-test-screenshot.png')
    val, left_top, _ = find_image_coordinates('./dev2/A-001-test-screenshot.png','./dev-img/dev2-chinese-header-resources-520x80.png')
    if val > 0.8:
        category = '提瓦特资源'
    else:
        val, left_top, _ = find_image_coordinates('./dev2/A-001-test-screenshot.png','./dev-img/dev2-chinese-header-Trophies-520x80.png')
        if val > 0.8:
            category = '掉落物'
        else:
            print('未知分类，退出')
            os._exit(0)
elif lag == 'english':
    screenshot_pic = py_screenshot('./dev2/A-001-test-screenshot.png')
    val, left_top, _ = find_image_coordinates('./dev2/A-001-test-screenshot.png','./dev-img/dev2-english-header-resources-520x80.png')
    if val > 0.8:
        category = "Teyvat's Resources"
    else:
        val, left_top, _ = find_image_coordinates('./dev2/A-001-test-screenshot.png','./dev-img/dev2-english-header-Trophies-520x80.png')
        if val > 0.8:
            category = 'Trophies'
        else:
            print('未知分类，退出')
            os._exit(0)
else:
    category = 'unknow'


"""
print('执行07步骤***************************************')
print('########### 07. 测试：对单个文字进行获取')
# 单击第一个，截图，再截取指定位置图片，然后对指定位置进行文字识别
py_click(140, 125, 114/2, 114/2)        # 步骤4，单击
screenshot_pic = py_screenshot('test-screenshot.png')    # B.步骤2，截图窗口并保存图片
crop_and_save_image(screenshot_pic, 1322, 666, 598, 228, 'test-crop_cnocr.png') # C.步骤3，对截图中进行文字截图，固定位置
textocr = text_ocr('test-crop_cnocr.png')     # 步骤6，对截图中部分进行文字识别并输出
print(textocr)

print('文字为',textocr[0]['text'],'分值为',textocr[0]['score'])
'''
文字位置,宽598，高228
左上    1322,666    高228
左下    1322,894                【下方文字位置为900高】
textocr的值为： print('文字为',textocr[0]['text'],'分值为',textocr[0]['score'])
[{'text': '铁块', 'score': 0.9998006224632263, 'position': array([[        462,         177],
       [        541,         177],
       [        541,         218],
       [        462,         218]], dtype=float32)}]
'''

# excel以追加的形式添加
from openpyxl import Workbook, load_workbook
import os
# 检查文件是否存在
filename = 'test.xlsx'
if not os.path.isfile(filename):
    # 创建一个新的工作簿
    workbook = Workbook()
    # 选择第一个工作表
    worksheet = workbook.active
    # 保存新建的 Excel 文件
    workbook.save(filename)
else:
    # 打开现有的 Excel 文件
    workbook = load_workbook(filename)
    # 选择第一个工作表
    worksheet = workbook.active
# 查找第一列中的空白行
column = worksheet['A']
empty_row = len(column) + 1
# 追加到第一列的空白行
worksheet[f'A{empty_row}'] = textocr[0]['text']
# 保存修改后的 Excel 文件
workbook.save(filename)
"""

print('执行07步骤***************************************')
print('########### 07. 实操：物产志中循环文字提取并输出图片')
from openpyxl import Workbook, load_workbook
import os
folder_name = "dev2"
if not os.path.exists(folder_name):
    # shutil.rmtree(folder_name)      # 清空和重建文件夹
    os.mkdir(folder_name)

examples_folder_name = "./dev2/examples"
if not os.path.exists(examples_folder_name):
    # shutil.rmtree(folder_name)      # 清空和重建文件夹
    os.mkdir(examples_folder_name)

filename = 'dev2/test.xlsx'
if not os.path.isfile(filename):
    # 创建一个新的工作簿
    workbook = Workbook()
    # 选择第一个工作表
    worksheet = workbook.active
    # 保存新建的 Excel 文件
    workbook.save(filename)

# A.单击，B.屏幕截图，C.对屏幕截图中的文字部分截图，D.获得文字，E.截图并将文字作为文件名保存选中的图片
# F.追加到excel，循环，G.横向5次，H.竖向5次，I.然后滚动，重复11次，J.对最后一部分进行额外判断  K.检测文字是否未变更，未变更退出 L.当值小于0.7时，添加excel行输出注意
######################  错误！！最后一行无法识别，且最后一部分会重复识别，造成图片错误，需要添加判断
column_x = 0  # 横向，列129
every_rows_y = 161    # 竖向，行161
out_for = False
last_word = ''
last_xy = False
scroll_first_png = f'./dev2/A-005-scroll_first_png.png'
every_rows = 4      # 重复得+1
top_left_y = 125 + window_top       # 首个Y坐标
rows = 0        # 行计数，不与循环一致，因为循环会重置
cols = 0        # 总数量
error_ocr = 0

# 判断当前为中文还是英文，物产志
# ./dev-img/dev2-english-210X75.png

if lag == 'english': 
    # lag = 'english' # 占前3排A/B/C 从标题来判断语言
    ex_a = 'A'
    ex_b = 'B'
    ex_c = 'C'
    ex_d = 'D'
elif lag == 'chinese':
    ex_a = 'E'
    ex_b = 'F'
    ex_c = 'G'
    ex_d = 'H'
else:
    # lag = 'other'
    ex_a = 'I'
    ex_b = 'J'
    ex_c = 'K'
    ex_d = 'L'

for h in range(1, 8 + 1):     # I.滚动，99次重复,1-9（8次）不含9
    if out_for == True:
        break
    if h > 1:
        isscroll = True
        py_scroll(every_rows)         # 步骤4，鼠标中键滚动,次数为滚动的行数
    else:
        isscroll = False
    # 循环5次，每次增加column_x
    for row in range(1, every_rows + 1):          # H.    每次识别4排/行 1-4
        rows += 1       # 行计数
        column_x = 0 # 换行后，横轴增值为0，重置
        if out_for == True:
            break
        # 对之前的内容循环5次，每次增加every_rows_y 161
        for jjj in range(1, 5 + 1):      # G. 1-5列(不含尾值)
            if out_for == True:
                break
            cols += 1    # 总数量
            if jjj == 1:    # 如果是第一列，用于判断最后一列第一个
                # crop_and_save_image(screenshot_pic, top_left_x, top_left_y, 114, 80, './dev2/A-006-last-check.png')    # 缩部分来验证///暂时用A-005-scroll_first_png.png来代替
                # last_one_first = ocr_png    # 对比文件名，暂时没用
                last_one_cols = cols        # 暂时没用，总数
                last_one_rows = rows        # 暂时没用，行数
            # 【仅限最后一排滚动后】该部分确定滚动后第一张图的位置进行定位
            if h > 1 and isscroll == True:     # 判断是否为滚动后首次，首次不相同。并且非未滚动前【非首次循环-需滚动后h==1，isscroll=h，确认是否滚动过】
                # J.最后一部分判断：记录上一部分的图片，滚动，识别核实上一张图片的位置，左上角>0.7,以新位置再次开始。如果未核实到，则按原顺序继续
                screen_png = py_screenshot('./dev2/A-003_check_screen_png.png',pic_save=True)               # 获得当前屏幕截图（软件中）
                # print('screen_png值为:',screen_png,'类型',type(screen_png))
                val, left_top, _ = find_image_coordinates('./dev2/A-003_check_screen_png.png','./dev2/A-004-check_test.png')               # ocr_png需要从上一次遍历时获取
                print('滚动后比对为',val,'坐标为：',left_top)
                if val > 0.8:   # 移动后比对的图片存在，且比对值大于0.8，说明已经到最后了，定位该地址,将其x坐标回归第一个，y坐标向下移动一类
                    print('滚动后比较图片时，发现相同图片')
                    left_sc = (140 + window_left, left_top[1] + 161 + window_top)     # 第一个坐标x=140，y+161，定位下一层的第一个。【需要额外接入偏移】
                    top_left_x = left_sc[0]
                    top_left_y = left_sc[1]
                    # 加入循环，然后替换掉top_left_x和top_left_y
                    print('滚动后重新比对坐标开始为：',top_left_x,top_left_y,'偏差为',window_left,window_top,'识别到的图片坐标为',left_top)
                    last_xy = True
                    for z in range(0,5):                # 向下滚动，保持最下方
                        pyautogui.scroll(-1,x=0,y=0)
                    time.sleep(0.5)
                else:
                    print('滚动后比较图片时，未发现相同图片，可能非最后一次对比','值为：',val,'坐标：',left_top)

            """
            该部分判断y轴实际值，分为首次【固定位置125+边框】、最后一次、
            """
            # 【Y轴坐标位置，分别为首次（固定）、最后一排、】
            top_left_x = 140 + column_x                                 # 横轴为固定值，首个为140
            if last_xy == False and h == 1 and row == 1:                             # 【第一次时/非最后一次、首行，Y轴为固定值】
                top_left_y = 125 + window_top
            elif last_xy == True:                                         # 【最后一排时】
                top_left_y = top_left_y + every_rows_y
            elif last_xy == False and isscroll == True and row == 1 and jjj == 1 :   # 【非最后一次，且滚动后,且第一次第一行第一排比对，确认滚动后第一个的坐标y】
                print('进入滚动后判定，进行比对')
                screenshot_pic = py_screenshot('./dev2/A-001-test-screenshot.png')
                val, left_top, _ = find_image_coordinates('./dev2/A-001-test-screenshot.png',scroll_first_png)
                if val < 0.8:
                    print('错误，滚动后未能发现首张图片位置')
                    os._exit(0)
                top_left_x = 140    # x坐标为首个，固定，column_x应该为0
                top_left_y = left_top[1]                                         # 【不变化】
            
            isscroll = False      # 改为非滚动后，以便后循环判断
            last_xy = False # 恢复判断
            click_loc = py_click(top_left_x, top_left_y , 114/2, 114/2)                                        # A.步骤4，单击，图片中间,后2个固定
            column_x += 129
            screenshot_pic = py_screenshot('./dev2/A-001-test-screenshot.png')                              # B.步骤2，截图窗口并保存图片
            crop_and_save_image(screenshot_pic, 1370, 830, 550, 65, './dev2/A-002-test-crop_cnocr.png')    # C.步骤3，对截图中进行文字截图，固定位置
            textocr = text_ocr('./dev2/A-002-test-crop_cnocr.png')                                      # D. 步骤6，对截图中部分进行文字识别并输出，textocr[0]['text']
            ocr_word = textocr[0]['text']
            ocr_score = round(textocr[0]['score'], 2)   # 2位小数
            import re
            invalid_chars = r'[\\/:*?"<>|-一]'  # 定义无效字符的正则表达式模式,\/:*?"<>|不能作为文件名，增加不允许添加的-
            if re.search(invalid_chars, ocr_word):
                ocr_word = re.sub(invalid_chars, '', ocr_word)
            """特别对待，该值会被重复造成中断，不需要考虑最后一行添加【过于相近】"""
            
            if "Artificed Spare Clockwork Component" in ocr_word:
                if error_ocr == 0:
                    ocr_word = "Artificed Spare Clockwork Component Coppelia"
                    error_ocr = 1
                else:
                    ocr_word = "Artificed Spare Clockwork Component Coppelius"
            ocr_png = f'./dev2/{lag}-{category}-{cols}-{rows}-{jjj}-{ocr_word}.png'        # 导出的图片
            shutil.copy2(f'./dev2/A-002-test-crop_cnocr.png','{examples_folder_name}/{lag}-{cols}-{ocr_word}.png')     # 复制文字部分的图片用于图像识别学习
            
            if last_word == ocr_word:               # K.检测文字是否未变更，未变更退出
                print('已到达最后，做最后一次叠加')
                click_loc = py_click(400, 400 , 114/2, 114/2)    # 在中间点以下
                # 把最后一行再来一次？
                for z in range(0,5):                # 向下滚动，保持最下方
                    pyautogui.scroll(-1,x=0,y=0)
                time.sleep(0.5)
                screen_png = py_screenshot('./dev2/A-003_check_screen_png.png',pic_save=True)
                val, left_top, _ = find_image_coordinates('./dev2/A-003_check_screen_png.png',scroll_first_png)
                if val > 0.8:   # 移动后比对的图片存在，且比对值大于0.8，说明已经到最后了，定位该地址,将其x坐标回归第一个，y坐标向下移动一类
                    print('滚动后比较图片时，发现相同图片')
                    left_sc = (140 + window_left, left_top[1] + window_top)     # 定位的是当前图片的位置
                    top_left_x = left_sc[0]
                    top_left_y = left_sc[1]
                    # 加入循环，然后替换掉top_left_x和top_left_y
                    print('滚动后重新比对坐标开始为：',top_left_x,top_left_y,'偏差为',window_left,window_top,'识别到的图片坐标为',left_top)
                    last_one_cols = last_one_cols - 1
                    column_x = 0
                for last_one in range(1, 5 + 1):
                    # print('开始最后一行')
                    last_one_cols += 1
                    top_left_x = 140 + column_x
                    click_loc = py_click(top_left_x, top_left_y , 114/2, 114/2)
                    screenshot_pic = py_screenshot('./dev2/A-001-test-screenshot.png')                              # B.步骤2，截图窗口并保存图片

                    crop_and_save_image(screenshot_pic, 1322, 666, 598, 228, './dev2/A-002-test-crop_cnocr.png')    # C.步骤3，对截图中进行文字截图，固定位置

                    textocr = text_ocr('./dev2/A-002-test-crop_cnocr.png')                                      # D. 步骤6，对截图中部分进行文字识别并输出，textocr[0]['text']
                    ocr_word = textocr[0]['text']
                    ocr_score = round(textocr[0]['score'], 2)   # 2位小数
                    import re
                    invalid_chars = r'[\\/:*?"<>|-]'  # 定义无效字符的正则表达式模式,\/:*?"<>|不能作为文件名，增加不允许添加的-
                    if re.search(invalid_chars, ocr_word):
                        ocr_word = re.sub(invalid_chars, '', ocr_word)
                    ocr_png = f'./dev2/{lag}-{category}-{last_one_cols}-{last_one_rows}-{last_one}-{ocr_word}.png'        # 导出的图片
                    shutil.copy2(f'./dev2/A-002-test-crop_cnocr.png','{examples_folder_name}/{lag}-{last_one_cols}-{ocr_word}.png')     # 复制
                    crop_and_save_image(screenshot_pic, top_left_x, top_left_y, 114, 114, ocr_png)  
                    column_x += 129
                    workbook = load_workbook(filename)
                    worksheet = workbook.active                 # 选择第一个工作表
                    column = worksheet[ex_a]                           # 查找第一列中的空白行
                    empty_row = len(column) + 1                         # 追加到第一列的空白行
                    worksheet[f'{ex_a}{empty_row}'] = ocr_word          # 第一列
                    worksheet[f'{ex_b}{empty_row}'] = ocr_score         # 第二列
                    worksheet[f'{ex_d}{empty_row}'] = category
                    if ocr_score < 0.7:
                        worksheet[f'{ex_c}{empty_row}'] = '该值可能存在问题，请验证'        # L.当值小于0.7时，添加excel行输出注意
                    # 保存修改后的 Excel 文件
                    workbook.save(filename)
                
                out_for = True
                break
            else:
                last_word = ocr_word
                
            # print('当前最后一张图片路径为：',ocr_png)
            crop_and_save_image(screenshot_pic, top_left_x, top_left_y, 114, 114, ocr_png)       # E.截图并将文字作为文件名保存选中的图片
            shutil.copy2(ocr_png,'./dev2/A-004-check_test.png')               # 将图片复制一份，保持英文，因为图像识别不支持中文文件名。会乱码
            
            """
            该部分为excel表格操作
            """
            # F.打开现有的 Excel 文件
            workbook = load_workbook(filename)
            worksheet = workbook.active                 # 选择第一个工作表
            column = worksheet[ex_a]                           # 查找第一列中的空白行
            empty_row = len(column) + 1                         # 追加到第一列的空白行
            worksheet[f'{ex_a}{empty_row}'] = ocr_word          # 第一列
            worksheet[f'{ex_b}{empty_row}'] = ocr_score         # 第二列
            worksheet[f'{ex_d}{empty_row}'] = category
            if ocr_score < 0.7:
                worksheet[f'{ex_c}{empty_row}'] = '该值可能存在问题，请验证'        # L.当值小于0.7时，添加excel行输出注意
            # 保存修改后的 Excel 文件
            workbook.save(filename)
            
            print('总数量:', cols ,'行', rows ,'列', jjj ,'识别文字：',f'{ocr_word}',' 图片名：',ocr_png,'识别分数：',ocr_score,'单击坐标：',click_loc,'滚动次数', h - 1,'次')
            if jjj == 5:
                print('-' * 30)
                
            if row == 4 and jjj == 5 :  # 在最后一列时，且循环4次后（要滚动前） 行4列5，指该部分的最后一个后=》滚动前
                top_left_x = 140
                top_left_y += every_rows_y
                print('滚动前首个Y轴',top_left_y)
                click_loc = py_click(top_left_x, top_left_y , 114/2, 114/2)
                screenshot_pic = py_screenshot('./dev2/A-001-test-screenshot.png')
                print('滚动前截取，截取坐标',top_left_x,top_left_y)
                crop_and_save_image(screenshot_pic, top_left_x, top_left_y, 114, 114, scroll_first_png)

            if jjj == 5:        # 最后一行不一定是最后一列，如果空白，则与前一个一致
                # print(top_left_y)
                top_left_y = top_left_y +  every_rows_y            # 首次、最后一行、滚动在后面有其他判断
                # print('值为：',top_left_y)       
                
    
    

    

print('分隔符**************************')



# step_x = 147 * 3    # 向右移动，背包147单个

# py_screenshot('test-screenshot.png')        # 步骤2，窗口截取截图
# crop_and_save_image(screenshot, 140, 125, 114, 114, 'test-cropped_image.png') # 步骤3，对截图中进行截图，物产志 140, 125, 114, 114 背包	118,121,124,124
# find_image_coordinates('test-screenshot.png','test-cropped_image.png')  # 步骤5，比对，image_path = 截图 和 template_path = 被比对的图  # 0.83

# py_click(140, 125, 114/2, 114/2)      # 步骤4，单击
# py_scroll(5)                          # 步骤4，鼠标中键滚动
# textocr = text_ocr('test-crop_cnocr.png')     # 步骤6，对截图中部分进行文字识别并输出


'''
待办：
1. 如果某个素材未获得，则可能中断【识别OCR空值】
'''