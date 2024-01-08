#
'''
创建时间：2024-01-02
修改时间：2024-01-02
该模块用于每次更新后，背包界面的图片、文字的截取进行遍历
'''
'''调用该库的模块

'''
from utils.ok_01_a_utils_uac import get_uac
from utils.ok_display_x_y import Display
import pygetwindow as gw
import pyautogui
import time
from PIL import Image
from PIL import ImageGrab
from multiprocessing import Process
import os
import keyboard
import cv2
import numpy as np
import sys
import shutil
from cnocr import CnOcr
from openpyxl import Workbook, load_workbook
from datetime import datetime
class CropPng:
    def __init__(self) -> None:
        '''
        - self.width 程序的宽 self.height 程序的高
        - self.window_left 程序距离屏幕左侧的距离 self.window_top 程序距离屏幕上部的距离
        '''
        self.process = None  # 实例变量，用于存储子进程
        # 01. 获得当前原神尺寸，并窗口置前
        dis = Display()
        self.width , self.height , self.window_left , self.window_top , _ = dis.windows_size_by_title('原神')
        # print(self.width , self.height , self.window_left , self.window_top)
        self.win_scale = self.width / 1920
        #print(win_scale)   #缩放比例，不是DPI，是跟1080P比较的比例 ，只支持16:9
        # xy坐标不含自身（不需要+1） 5次滚动一循环
        self.left_top_x = int(117 * self.win_scale)                   # 第一排第一个x轴[随尺寸变化]1080P下
        self.col_x = int(147 * self.win_scale)                    # 素材图片每排向右偏移147，每行有8个[随尺寸变化]1080P下
        self.row_y = int(175  * self.win_scale)                    # 素材图片每行向下偏移175，单屏幕有4行有效[随尺寸变化]1080P下
        ''' +124到数字
        y0第1行121,1数245 , 第2行296,2数420 , 第3行471,3数595 , 第4行646,2数770
        y1第1行124,1数248 , 第2行299,2数423 , 第3行474,3数598 , 第4行649,2数773
        y2第1行110,1数234 , 第2行285,2数409 , 第3行460,3数584 , 第4行635,2数759
        y3第1行113,1数237 , 第2行288,2数412 , 第3行463,3数587 , 第4行638,2数762
        y4第1行117,1数241 , 第2行292,2数416 , 第3行467,3数591 , 第4行642,2数766
        '''
        self.num_y = int(128 * self.win_scale)
        self.num_y_size = int(24 * self.win_scale)
        # 第0、5、10、15次滚动
        self.left_top_y0_l1 = int(121 * self.win_scale )    # 0              # [未滚动0或5、10、15]第一排第一个y轴，竖向移动175[随尺寸变化]1080P下
        self.left_top_y0_l2 = int(self.left_top_y0_l1 + self.row_y * 1)
        self.left_top_y0_l3 = int(self.left_top_y0_l1 + self.row_y * 2)
        self.left_top_y0_l4 = int(self.left_top_y0_l1 + self.row_y * 3)
        self.left_top_y0_n1 = int(self.left_top_y0_l1 + self.num_y + self.row_y * 0)     # 第1行数字y
        self.left_top_y0_n2 = int(self.left_top_y0_l1 + self.num_y + self.row_y * 1)     # 第2行数字y
        self.left_top_y0_n3 = int(self.left_top_y0_l1 + self.num_y + self.row_y * 2)
        # 第1、6、11、16次滚动
        self.left_top_y1_l1 = int(124 * self.win_scale)     # +3             # [第1、6、11次滚动]第一排第一个y轴，竖向移动175[随尺寸变化]1080P下
        self.left_top_y1_l2 = int(self.left_top_y1_l1 + self.row_y * 1)
        self.left_top_y1_l3 = int(self.left_top_y1_l1 + self.row_y * 2)
        self.left_top_y1_l4 = int(self.left_top_y1_l1 + self.row_y * 3)
        self.left_top_y1_n1 = int(self.left_top_y1_l1 + self.num_y + self.row_y * 0)
        self.left_top_y1_n2 = int(self.left_top_y1_l1 + self.num_y + self.row_y * 1)
        self.left_top_y1_n3 = int(self.left_top_y1_l1 + self.num_y + self.row_y * 2)
        # 第2、7、12、17次滚动
        self.left_top_y2_l1 = int(110 * self.win_scale)     # -11             # [第2、7、12次滚动]第一排第一个y轴，竖向移动175[随尺寸变化]1080P下 额外一次移动
        self.left_top_y2_l2 = int(self.left_top_y2_l1 + self.row_y * 1)
        self.left_top_y2_l3 = int(self.left_top_y2_l1 + self.row_y * 2)
        self.left_top_y2_l4 = int(self.left_top_y2_l1 + self.row_y * 3)
        self.left_top_y2_n1 = int(self.left_top_y2_l1 + self.num_y + self.row_y * 0)
        self.left_top_y2_n2 = int(self.left_top_y2_l1 + self.num_y + self.row_y * 1)
        self.left_top_y2_n3 = int(self.left_top_y2_l1 + self.num_y + self.row_y * 2)
        # 第3、8、13、18次滚动
        self.left_top_y3_l1 = int(113 * self.win_scale)     # -8             # [第3、8、13次滚动]第一排第一个y轴，竖向移动175[随尺寸变化]1080P下
        self.left_top_y3_l2 = int(self.left_top_y3_l1 + self.row_y * 1)
        self.left_top_y3_l3 = int(self.left_top_y3_l1 + self.row_y * 2)
        self.left_top_y3_l4 = int(self.left_top_y3_l1 + self.row_y * 3)
        self.left_top_y3_n1 = int(self.left_top_y3_l1 + self.num_y + self.row_y * 0)
        self.left_top_y3_n2 = int(self.left_top_y3_l1 + self.num_y + self.row_y * 1)
        self.left_top_y3_n3 = int(self.left_top_y3_l1 + self.num_y + self.row_y * 2)
        # 第4、9、14、19次滚动
        self.left_top_y4_l1 = int(117 * self.win_scale)     # -4              # [第4、9、14次滚动]第一排第一个y轴，竖向移动175[随尺寸变化]1080P下
        self.left_top_y4_l2 = int(self.left_top_y4_l1 + self.row_y * 1)
        self.left_top_y4_l3 = int(self.left_top_y4_l1 + self.row_y * 2)
        self.left_top_y4_l4 = int(self.left_top_y4_l1 + self.row_y * 3)
        self.left_top_y4_n1 = int(self.left_top_y4_l1 + self.num_y + self.row_y * 0)
        self.left_top_y4_n2 = int(self.left_top_y4_l1 + self.num_y + self.row_y * 1)
        self.left_top_y4_n3 = int(self.left_top_y4_l1 + self.num_y + self.row_y * 2)
        self.rect_xy = int(124  * self.win_scale )                # 正方形，素材图片的大小 除以2就是居中了[随尺寸变化]1080P下
        self.rect_xy_half = int(self.rect_xy / 2)    # 一半的方形 =62
        self.safe_bottom_y = 0
        # 滚动一次16px

    def py_screenshot(self, screenshot_png = None, pic_save=False,mult_pro=False):      # mult_pro多线程保存
        '''游戏内，去掉了屏幕边界部分
        如果保存，则以1080P保存，不保存则为原尺寸
        - screenshot_png 保存的路径，如果不保存则None
        - pic_save 是否保存为图片
        - mult_pro 保存前提下是否以多线程来保存
        '''
        screenshot_pic = pyautogui.screenshot(region=(self.window_left, self.window_top, self.width, self.height))  # 游戏内的截图，不含其他

        if pic_save:
            pic_width, pic_height = screenshot_pic.size
            #print('截图尺寸',pic_width,pic_height)
            resized_screenshot = screenshot_pic.resize((1920, 1080))
            if mult_pro:
                # 创建一个新的子进程，执行所有操作。如果主进程过短，会造成无法保存
                self.process = Process(target=self.save_screenshot_in_process, args=(resized_screenshot, screenshot_png),daemon=True)
                self.process.start()
                return self.process
            else:
                self.save_screenshot_in_process(resized_screenshot, screenshot_png)
        else:
            return screenshot_pic   # 

    def save_screenshot_in_process(self,resized_screenshot, screenshot_png):
        try:    # 子进程try，防止退出错误，造成崩溃
            resized_screenshot.save(screenshot_png)
        except:
            print('发送错误')
            exit(1)
    def crop_and_save_image(self,screenshot, x, y, width, height, save_path):        # 从前端获得A. screenshot截图，B C.定义x、y坐标，D E.宽高矩形，F. 保存位置
        '''对图片或截图再次进行截取
        '''
        # 判断输入类型是图像路径还是图像对象
        if isinstance(screenshot, str):  # 如果是字符串，则加载图像
            screenshot = Image.open(screenshot)
        # 根据给定的坐标和尺寸进行截取
        cropped_image = screenshot.crop((x, y, x + width, y + height))
        # 保存截取的图片
        cropped_image.save(save_path)
    # crop_and_save_image(screenshot, 140, 125, 114, 114, 'test-cropped_image.png')

    def py_click(self,x,y,width,height):     # x坐标，y坐标=左上角顶点坐标，width、height=宽高中心点增值，也可以设为0
        '''对指定位置进行单击操作，x、y需要尺寸变化、去掉边界距离部分
        '''
        click_loc = (x + width + self.window_left,y + height + self.window_top)    # 左上角顶点+中心点增值 + 窗口与屏幕的差距
        pyautogui.click(click_loc) 
        pyautogui.click(click_loc)  # 单击两下，以便去掉【新】字
        # time.sleep(0.5)             # 等待一下，以便让图片能够动画完成
        return click_loc
    # py_click(140, 125, 114/2, 114/2)


    def find_image_coordinates(self,image_path, template_path):      # image_path = 屏幕截图 template_path = 对比图,添加了，输入内容判断是图片还是图片的路径
        # 判断image_path和template_path是否为字符串路径
        if isinstance(image_path, str):
            # 读取图像文件
            image = cv2.imread(image_path)
        else:
            # image_path为图像变量
            image = np.array(image_path)
            image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

        if isinstance(template_path, str):
            # 读取模板文件
            template = cv2.imread(template_path)
        else:
            # template_path为图像变量
            template = np.array(template_path)
            template = cv2.cvtColor(template, cv2.COLOR_RGB2BGR)
        # 使用模板匹配算法查找模板在图像中的位置
        result = cv2.matchTemplate(image, template, cv2.TM_CCOEFF_NORMED)
        _, max_val, _, max_loc = cv2.minMaxLoc(result)      # max_val是相似度，max_loc为坐标（元组）
        # 获取模板的宽度和高度
        template_width = template.shape[1]
        template_height = template.shape[0]

        # 计算模板的左上角和右下角坐标
        top_left = max_loc      # 左上角坐标
        bottom_right = (top_left[0] + template_width, top_left[1] + template_height)    # 右下角坐标，依赖比对图的大小，如果有差异、拉伸则错误，114或125
        #print('比对相似度为：',max_val, '坐标为：',top_left, bottom_right)
        return max_val,top_left, bottom_right       # 相似度,左上角坐标，右下角坐标

    def text_ocr(self,pic_ocr,single=False):
        ocr = CnOcr(det_model_name='naive_det') 
        if single == True:
            text = ocr.ocr_for_single_line(pic_ocr) 
        else:
            text = ocr.ocr(pic_ocr) 
        return text



    # 01. 循环以截取整个截图并保存
    '''背包 -养成道具、材料
    A标题文字定位：养成道具、材料：(130,30) 250x40
    B一排1：(118,121)  123x123
    C一排2：(265,121)   左右偏移【146】
    D二排1：(118,296)   上下偏移【175】
    一排8个 +147  最后一个1147
    一个界面4+1（第5显示一半）
    E文字-右1【名称】-大概率单行：【多行：奇械发条备件·科培琉司】 (1315,127) 482x50 (字可能会侵入下方)
    F文字-右2【分类】-单行：(1315,180) 265x40
    G数字-【数量】-单行：(118,244)124x25  Y轴+123
    '''
    """
    1.A【单行】对标题(130,30) 250x40部分进行判断，确定类目
    2.E【单行】名称（右1文字识别）(1334,127)  (部分有多行)、截图（440x45）用于OCR学习 命名：A-总数-排数-列数-名称-分类.png【英文有2行，y上+5，下+6=》未加入】
    3.F【单行】分类（右2文字识别）(1334,180) 、截图（245x40）用于OCR学习 ，只有4类？ 命名：B-总数-排数-列数-名称-分类.png【英文有2行，要y+15=》未加入】
    4.G【单行】数量（数字识别） x+10,Y轴+128 ，(128,249)截图（100x25），用于OCR学习，该值由于Y轴坐标差异，有偏移
    5.B 首位定位(118,121)，A.截图(123x123) 命名：总数-排数-列数-名称-分类.png
    6.加入到excel表格中
    循环2-5，一行8个，3行，+1行，滚动3行，重新定位
    990.X轴循环，7次（8个），每次+147 column_x
    991.Y轴循环，3排+第一个截图
    992.滚动，循环，滚动，重新截图-定位
    993.最后一排判断
    """
    def loop_to_crop(self):
        scoll = 99              # 设定的滚动次数

    def test(self):
        x = self.left_top_x
        y = self.left_top_y
        nums = 1        # 总数量
        rows = 1        # 行数
        cols = 1        # 列数
        folder_name = "img"
        for _ in range(1,4 + 1):    # 进行行，y轴移动
            
            for _ in range(1 , 8 + 1):  # 第一行，8个，横向点击列
                self.py_click( x , y , self.rect_xy / 2 , self.rect_xy / 2)   # 单击第一排第一个素材图片的位置，118,121   居中123/123
                x += self.col_x
                self.py_screenshot(f'{folder_name}\\{rows}-{cols}-{nums}.png',pic_save=True)
                nums += 1
                cols += 1
            
            rows += 1 
            cols = 1    # 重置排序列
            x = self.left_top_x     # 重置x轴定位
            y += self.row_y         # y轴叠加距离
            
            # 到底了，额外步骤，
         
    # 模拟鼠标按住、拖动、释放操作的示例代码
    def simulate_drag_and_drop(self,start_x, start_y, end_x, end_y, duration=1):        # 游戏内移动，去掉了屏幕边界
        # 移动到起始位置
        pyautogui.moveTo(start_x +self.window_left, start_y +self.window_top)
        #time.sleep(5)
        # 鼠标按住
        pyautogui.mouseDown()
        #time.sleep(1)
        # 使用 moveTo 函数进行拖动（平滑移动）
        pyautogui.moveTo(end_x + self.window_left, end_y  + self.window_top, duration=duration)
        time.sleep(1)
        # 鼠标释放
        pyautogui.mouseUp()
            
    def test2(self):
        
        x = self.left_top_x + self.rect_xy / 2
        y = self.left_top_y + self.rect_xy / 2 - 35 * self.win_scale        # 偏移25（1080P）测试4k未测试1080P
        y2 = self.left_top_y + self.rect_xy / 2 + self.row_y * 4
        print('x',x,'y',y,'y2',y2)
        for _ in range(8):
            self.simulate_drag_and_drop(x,y2,x,y)
            time.sleep(2)       # 拖动等待
    def test3(self):
        image_path = f'pic\\test3.png'    # 屏幕截图
        template_path = f'pic\\down.png'
        self.find_image_coordinates(image_path, template_path)


    def test01(self):
        # 对原神进行截图
        folder_name = 'test'
        #self.py_screenshot(f'{folder_name}\\screenshot.png',pic_save=True)
        self.py_screenshot(f'{folder_name}\\last_row_09_摆设.png',pic_save=True)
        

    def test02(self):
        x = self.left_top_x
        rect_xy = self.rect_xy          # 方块124
        rect_half = self.rect_xy_half   # 方块一半，62
        filename = f'./test/test.xlsx'
        if os.path.isfile(filename):
            os.remove(filename)
        workbook = Workbook()
        worksheet = workbook.active
        workbook.save(filename)
        
        # 第1次滚动
        for i in range(0,10 +1):
            index = i % 5       # 取余数 获得y坐标
            if index == 0:
                y1 =  self.left_top_y0_l1   # 左上角位置
                y2 = self.left_top_y0_l2
                y3 = self.left_top_y0_l3
                y4 = self.left_top_y0_l4
                y_n1 = self.left_top_y0_n1  # 第一行数字
                y_n2 = self.left_top_y0_n2
                y_n3 = self.left_top_y0_n3
            elif index == 1:
                y1 =  self.left_top_y1_l1   # 左上角位置
                y2 = self.left_top_y1_l2
                y3 = self.left_top_y1_l3
                y4 = self.left_top_y1_l4
                y_n1 = self.left_top_y1_n1  # 第一行数字
                y_n2 = self.left_top_y1_n2
                y_n3 = self.left_top_y1_n3
            elif index == 2:
                y1 =  self.left_top_y2_l1   # 左上角位置
                y2 = self.left_top_y2_l2
                y3 = self.left_top_y2_l3
                y4 = self.left_top_y2_l4
                y_n1 = self.left_top_y2_n1  # 第一行数字
                y_n2 = self.left_top_y2_n2
                y_n3 = self.left_top_y2_n3
            elif index == 3:
                y1 =  self.left_top_y3_l1   # 左上角位置
                y2 = self.left_top_y3_l2
                y3 = self.left_top_y3_l3
                y4 = self.left_top_y3_l4
                y_n1 = self.left_top_y3_n1  # 第一行数字
                y_n2 = self.left_top_y3_n2
                y_n3 = self.left_top_y3_n3
            elif index == 4:
                y1 =  self.left_top_y4_l1   # 左上角位置
                y2 = self.left_top_y4_l2
                y3 = self.left_top_y4_l3
                y4 = self.left_top_y4_l4
                y_n1 = self.left_top_y4_n1  # 第一行数字
                y_n2 = self.left_top_y4_n2
                y_n3 = self.left_top_y4_n3
            
            def loop_col(last_row = False,last_row_num = 0):
                if last_row_num == 4:
                    nnn = 4
                else:
                    nnn = 3
                for ii in range(1, nnn +1):    # 3行 或4行
                    xx = self.left_top_x    # 重置x轴
                    if ii == 1:
                        y = y1
                        y_n = y_n1
                    elif ii ==2:
                        y = y2
                        y_n = y_n2
                    elif ii ==3:
                        y = y3
                        y_n = y_n3
                    elif ii == 4:
                        y = y4
                        y_n = y_n4
                    for i in range(1, 8 + 1):   # 8列
                        # 01.单击方框
                        self.py_click(xx , y , rect_half , rect_half)
                        # 02.获得右侧文字-子进程 (1334,127)  2636，254
                        right_word_1_x = 1334 * self.win_scale
                        right_word_1_y = 127 * self.win_scale
                        right_word_1_xx = 440 * self.win_scale
                        right_word_1_yy = 45 * self.win_scale
                        
                        if last_row:
                            
                            screen_pic = self.py_screenshot(None,pic_save = False,mult_pro=False)   # 不保存为原始尺寸截图
                            self.crop_and_save_image(screen_pic , right_word_1_x, right_word_1_y ,right_word_1_xx , right_word_1_yy , f'test\\right_word_1.png')   
                            # 文字获得
                            textocr = self.text_ocr(f'test\\right_word_1.png',single=True)
                            right_word_name = textocr['text']
                            # 测试右侧文字是否一致
                            if not i == 1:
                                if comp_word == right_word_name:
                                    print('右侧文字识别一致，退出')
                                    end_time = time.time()
                                    execution_time = end_time - start_time
                                    minutes = int(execution_time // 60)
                                    seconds = int(execution_time % 60)
                                    print(f"代码执行时间: {minutes} 分钟 {seconds} 秒")
                                    current_time = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
                                    new_filename = f'./test/test-{current_time}.xlsx'
                                    shutil.copy(filename, new_filename)
                                    sys.exit(0)
                            comp_word = right_word_name
                            
                            # 02-1.备份右侧图片区域
                            shutil.copy(f'test\\right_word_1.png', f'test\\right_word_0.png')
                        else:
                            screen_pic = self.py_screenshot(None,pic_save = False,mult_pro=False)   # 不保存为原始尺寸截图
                            self.crop_and_save_image(screen_pic , right_word_1_x, right_word_1_y ,right_word_1_xx , right_word_1_yy , f'test\\right_word_1.png')   
                        # 文字获得
                        textocr = self.text_ocr(f'test\\right_word_1.png',single=True)
                        right_word_name = textocr['text']
                        
                        # 03.获得分类-子进程 (1334,180)
                        right_word_2_x = 1334 * self.win_scale
                        right_word_2_y = 180 * self.win_scale
                        right_word_2_xx = 245 * self.win_scale
                        right_word_2_yy = 40 * self.win_scale
                        self.crop_and_save_image(screen_pic , right_word_2_x, right_word_2_y ,right_word_2_xx , right_word_2_yy , f'test\\right_word_2.png')   
                        textocr = self.text_ocr(f'test\\right_word_2.png',single=True)
                        right_word_name_2 = textocr['text']
                        
                        # 04.获得数字-子进程
                        self.crop_and_save_image(screen_pic , xx, y_n ,self.rect_xy , self.num_y_size , f'test\\num_1.png')   
                        textocr = self.text_ocr(f'test\\num_1.png',single=True)
                        numnum = textocr['text']
                        # 05.x+1
                        workbook = load_workbook(filename)
                        worksheet = workbook.active
                        column = worksheet['A'] 
                        empty_row = len(column) + 1
                        worksheet[f'B{empty_row}'] = right_word_name_2
                        worksheet[f'C{empty_row}'] = right_word_name
                        worksheet[f'D{empty_row}'] = numnum
                        workbook.save(filename)
                        print(f'{right_word_name_2}, {right_word_name}, {numnum}')
                        print('*' * 20)
                        
                        xx += self.col_x

            
            '''
            01. 点击y1
            07. [非第一回合]与04对比，确定y1值与06是否一致 不一致则为最后一次，不再滚动，判断最后行数
            02. 顺序3排【如果右侧一致，则为最后一个】
            03. 点击y4
            04. 截图y4
            05. 滚动29 +1
            --- 下一回合
            
            '''
            # 测试，滚动前的屏幕截图
            #self.py_click(x , y4 , rect_half , rect_half)
            #screen_pic = self.py_screenshot(f'test\\第{i}次滚动前.png',pic_save = True,mult_pro=False)
            
            # 01. 点击y1
            self.py_click(x , y1 , rect_half , rect_half)
            time.sleep(0.2)
            # 07. [非第一回合]与04对比，确定y1值与06是否一致 不一致则为最后一次，不再滚动，判断最后行数
            if not i == 0:
                screen_pic = self.py_screenshot(None,pic_save = False,mult_pro=False)   # 不保存为原始尺寸截图
                max_val,top_left, _ = self.find_image_coordinates(screen_pic,f'test\\crop_{i -1}.png')      #原始截图进行对比
                print(f'{i}，y1值为：{y1} ，识别y:{top_left[1]} ,差值为:{top_left[1] - y1}')
                # 还有少于第一次的情况
                if abs(top_left[1] - y1) > 10:
                    if top_left[1] - y4 > 0:    # 在第四排下面
                        '''
                        @@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@这里要重新计算Y坐标，同时给方法传递个参数确定行数和最后一次'''
                        print('最后还有1行，y4：',{y4})
                        
                        y1 = y1 + top_left[1] - y4
                        y2 = y1 + self.row_y * 1
                        y3 = y1 + self.row_y * 2
                        y_n1 = y1 + self.num_y + self.row_y * 0
                        y_n2 = y1 + self.num_y + self.row_y * 1
                        y_n3 = y1 + self.num_y + self.row_y * 2
                        print(f'更新后，y1:{y1} ,y2:{y2} ,y3:{y3} ,数字y_n1:{y_n1} ,y_n2:{y_n2} ,y_n3:{y_n3}')
                        loop_col(last_row = True,last_row_num = 1)
                    elif top_left[1] - y3 > 0:
                        print('最后还有2行，y3：',{y3})
                        y1 = y1 + top_left[1] - y3
                        y2 = y1 + self.row_y * 1
                        y3 = y1 + self.row_y * 2
                        y_n1 = y1 + self.num_y + self.row_y * 0
                        y_n2 = y1 + self.num_y + self.row_y * 1
                        y_n3 = y1 + self.num_y + self.row_y * 2
                        print(f'更新后，y1:{y1} ,y2:{y2} ,y3:{y3} ,数字y_n1:{y_n1} ,y_n2:{y_n2} ,y_n3:{y_n3}')
                        loop_col(last_row = True,last_row_num = 2)
                    elif top_left[1] - y2 > 0:
                        print('最后还有3行，y2：',{y2})
                        y1 = y1 + top_left[1] - y2
                        y2 = y1 + self.row_y * 1
                        y3 = y1 + self.row_y * 2
                        y_n1 = y1 + self.num_y + self.row_y * 0
                        y_n2 = y1 + self.num_y + self.row_y * 1
                        y_n3 = y1 + self.num_y + self.row_y * 2
                        print(f'更新后，y1:{y1} ,y2:{y2} ,y3:{y3} ,数字y_n1:{y_n1} ,y_n2:{y_n2} ,y_n3:{y_n3}')
                        loop_col(last_row = True,last_row_num = 3)
                    elif top_left[1] - y1 > 0:
                        print('最后还有4行，y1：',{y1})
                        y1 = y1 + top_left[1] - y1
                        y2 = y1 + self.row_y * 1
                        y3 = y1 + self.row_y * 2
                        y4 = y1 + self.row_y * 3
                        y_n1 = y1 + self.num_y + self.row_y * 0
                        y_n2 = y1 + self.num_y + self.row_y * 1
                        y_n3 = y1 + self.num_y + self.row_y * 2
                        y_n4 = y1 + self.num_y + self.row_y * 3
                        print(f'更新后，y1:{y1} ,y2:{y2} ,y3:{y3} ,数字y_n1:{y_n1} ,y_n2:{y_n2} ,y_n3:{y_n3}')
                        loop_col(last_row = True,last_row_num = 4)
                    else:
                        print('错误')
                        end_time = time.time()
                        execution_time = end_time - start_time
                        minutes = int(execution_time // 60)
                        seconds = int(execution_time % 60)
                        print(f"代码执行时间: {minutes} 分钟 {seconds} 秒")
                        sys.exit()
            else:
                print(f'{i},1值为：{y1}')
            # 02. 顺序3排【如果右侧一致，则为最后一个】
            loop_col()
            # 03. 点击y4
            self.py_click(x , y4 , rect_half , rect_half)      # 单击第四排第一个位置
            time.sleep(0.2)
            # 04. 截图y4
            screen_pic = self.py_screenshot(None,pic_save = False,mult_pro=False)   # 不保存为原始尺寸截图
            self.crop_and_save_image(screen_pic , x, y4 ,rect_xy , rect_half , f'test\\crop_{i}.png')   # 滚动前的截图，截图为原始尺寸 ，第四行第一排的一半
            # 05. 滚动29 +1
            for _ in range(29):
                pyautogui.scroll(-1, x=0, y=0)  # 滚动一次16px 4k[24]
            if index == 1:    # 第2\7\12次
                print('额外向下',i)
                pyautogui.scroll(-1, x=0, y=0)
            time.sleep(0.2)
            
        
    def test03(self):   # 输出指定图片指定坐标的颜色

        # 加载图片
        img = Image.open('1.png')

        # 获取坐标(1, 2)的像素颜色，注意在PIL中坐标是(x, y)，从左上角开始计数
        color = img.getpixel((117, 291))

        # 输出颜色，通常是一个RGB元组
        print(color)
    
    def test04(self):
        for i in range(1,20+1):
            index = i % 5
            print(i,'取余数',index)
            
    def test05(self):
        a = self.win_scale
        y0 = self.left_top_y0
        y1 = self.left_top_y1
        y2 = self.left_top_y2
        y3 = self.left_top_y3
        y4 = self.left_top_y4
        z1 = 0
        z2 = 175 * 1
        z3 = 175 * 2
        z4 = 175 * 3
        n = 124
        print(f'y0第1行{y0},1数{y0 + z1 + n} , 第2行{y0 + z2},2数{y0 + z2 + n} , 第3行{y0 + z3},3数{y0 + z3 + n} , 第4行{y0 + z4},2数{y0 + z4 + n}')
        print(f'y1第1行{y1},1数{y1 + z1 + n} , 第2行{y1 + z2},2数{y1 + z2 + n} , 第3行{y1 + z3},3数{y1 + z3 + n} , 第4行{y1 + z4},2数{y1 + z4 + n}')
        print(f'y2第1行{y2},1数{y2 + z1 + n} , 第2行{y2 + z2},2数{y2 + z2 + n} , 第3行{y2 + z3},3数{y2 + z3 + n} , 第4行{y2 + z4},2数{y2 + z4 + n}')
        print(f'y3第1行{y3},1数{y3 + z1 + n} , 第2行{y3 + z2},2数{y3 + z2 + n} , 第3行{y3 + z3},3数{y3 + z3 + n} , 第4行{y3 + z4},2数{y3 + z4 + n}')
        print(f'y4第1行{y4},1数{y4 + z1 + n} , 第2行{y4 + z2},2数{y4 + z2 + n} , 第3行{y4 + z3},3数{y4 + z3 + n} , 第4行{y4 + z4},2数{y4 + z4 + n}')
    '''
    第一列图片x 左侧距离117  
    上部y 96px
    第一行y距离上部 25px
    图片 内宽度x 123
    图片 内高度y 123+29（数字）
    图片y 1-2行175
    第1-2行y中间 23px
    列之间x 23px 从左上角146
    
    
    第一行  【1】x横 = 117 y高 = 96+25 =117  【2】 x = 117+ 
    '''





if __name__ == "__main__":
    # 01. 获得管理员权限
    get_uac()
    
    
    # 输出开始时间
    start_time = time.time()
    formatted_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(start_time))
    print("开始时间:", formatted_time)
    crop = CropPng()
    print('游戏分辨率为',crop.width,crop.height)
    
    def on_key_release(event):
        if event.name == 'f3':
            # 按下 F3 键后终止程序
            print("按下 F3 键，程序退出。")
            end_time = time.time()
            execution_time = end_time - start_time
            minutes = int(execution_time // 60)
            seconds = int(execution_time % 60)
            print(f"代码执行时间: {minutes} 分钟 {seconds} 秒")

            os._exit(0)

    # 注册按键释放事件的回调函数
    keyboard.on_release(on_key_release)
    
    # test01 对原神界面进行屏幕截图
    #crop.test01()
    # test02 滚动
    crop.test02()
    #crop.test03()
    #crop.test04()
    #crop.test05()
    
    time.sleep(3)   # 等待，使得截图正确截取原神
    #crop.py_screenshot('test-screenshot.png')  # 直接的截图
    #crop.test()
    #crop.test2()
    
    
    
    
    
    # 获得排的第一个区域图片，如果非指定Y轴，则为尾部，然后根据Y轴位置来判断可能存在的剩余行数
    # 第一行距离上部50，第二行为45，建议40（4k下）
    '''
    识别第二行应该的y轴（x轴固定）选中第一个的状态
    
    1.点击第一个区域
    2.截图
    3.上拉
    4.判断截图的位置 与 定义的第一排位置相等，相等则直接下一步
        -不相等，判断截图的y坐标是否为第一排的区域
            - 在，继续拉
            - 不再，为最后一部分了，按新位置
                - 根据当前位置，判断还应该有多少行
                    - 在最后一行时，判断右侧是否一致，如一致，则不保存图片
                    - 如果正好为第八列，则终止
    
    a的目的：获得素材部分的截图和名称 =》 一版本一次
    b的目的：手动更新数值，去掉a的x轴变动，进行大范围的截图。然后根据素材截图获得定位，获得素材的数值导出到excel
    
    '''