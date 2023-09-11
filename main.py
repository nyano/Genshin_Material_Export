"""
0.游戏置顶
1.识别指定图片，然后OCR指定位置的数字，将数字写入文件
2.翻页
3.保存在excel

"""
import cv2
from PIL import Image
import os
import pyautogui
import numpy as np
import time
from find_process import get_window_handle
# opencv不能有中文文件名
from process_x_y import get_width

# '''
if not get_window_handle('原神'):   # 将标题名为原神的进程前置
    print('你都没开原神！')
    os._exit(1)

real_width , real_height ,scaling ,borderless = get_width('原神')
if not real_width == 1920 or not real_height == 1080:
    print('窗口不是1920X1080')
    os._exit(1)
# '''

# 通用部分
import configparser
# 读取配置文件
config = configparser.ConfigParser()
config.read('config.ini', encoding='utf-8')
image_path = config['lag']['image_path']
sheetname = config['lag']['sheetname']

# 创建保存截图的目录

screenshot_folder = './screenshot'

if not os.path.exists(screenshot_folder):
   os.makedirs(screenshot_folder)

for filename in os.listdir(screenshot_folder):
   file_path = os.path.join(screenshot_folder, filename)
   if os.path.isfile(file_path):
       os.remove(file_path)

# 1.首先分割文件名将存入列表
# 【1-big_experience.png】分割，获得big_experience,按顺序获得num_list
# [[num_list,big_experience,1-big_experience.png]]  [[排序数字，道具名，图片名]]


img_list = []
num = 0

# 遍历image_path目录下的文件
for filename in os.listdir(image_path):
    if filename.endswith('.png'):
        # 提取文件名中的数字部分
        num += 1
        
        # 提取文件名中的标识部分
        identifier = filename.split('-')[1].split('.')[0]
        
        # 将文件名、标识和完整文件名添加到img_list中
        img_list.append([num, identifier, filename])


# 根据num_list的顺序对img_list进行排序
sorted_img_list = sorted(img_list, key=lambda x: x[0])
print(sorted_img_list)
# sorted_img_list = [[1, 'big_experience', '1-big_experience.png'], [2, 'mid_experience', '2-mid_experience.png'], [3, 'zzzzzzzz', '3-zzzzzzzz.png']]
# 打印排序后的结果
for item in sorted_img_list:
    print(item)
    
print('*'*20)

list_num = 0 
first_loc = None
error_num = 0
rows = 0    # 行数
error_loc = None
while list_num < num:
    print('#' * 20)
    print('错误循环次数：',error_num,'当前次数',list_num,'当前行',rows)
    list_num += 1
    matching_list = [sub_list for sub_list in sorted_img_list if sub_list[0] == list_num]
    print('当前列表',matching_list)
    list_a = matching_list[0][0]        # 序列，用于定位
    list_b = matching_list[0][1]        # 道具名
    list_c = matching_list[0][2]        # 图片名
    print('列表参数为',list_a,list_b,list_c)
    # 构建完整的图片路径
    image_path_file = os.path.join(image_path, list_c)
    
    # 加载当前图片
    current_image = cv2.imread(image_path_file) 
    # 加载当前屏幕截图
    screenshot = pyautogui.screenshot()
    screenshot = np.array(screenshot)
    screenshot = cv2.cvtColor(screenshot, cv2.COLOR_RGB2BGR)
    result = cv2.matchTemplate(current_image, screenshot, cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    print("匹配图为", list_c, "匹配值为", max_val, "匹配坐标为", max_loc)
    print(list_num)
    if list_num == 1:           # 仅限第一行
        first_loc = max_loc
    if rows > 0 :       # 不是第一行的时候
        if abs(error_loc[1] - max_loc[1]) <= 3:    # 判断是否在同一高度，如果在，说明未能拉高度  +-3范围内，一定偏差
            print('未能拉高度')             # 跳过判断
            break
        else:
            first_loc = max_loc     # 定位首行坐标
            rows -= 1
            error_loc = None
    # 检查 max_val 的 y 值减去 first_loc 的 y 值是否大于 630，并输出结果
    # 单行165【点对点180】,4行680
    # 如果首次就在下方，则不被识别
    print('当前第一次坐标',first_loc,'当前坐标',max_loc)
    if max_loc[1] - first_loc[1] > 165:     # 高度对比，如果处于第二行，则进行鼠标移动
        error_loc = max_loc
        list_num -= 1 # 重来一次当前的循环
        print('返回次数',list_num)
        rows += 1   # 识别自己不是第一行了
        error_num += 1
        if error_num >= 3:      # 重复次数大于3次，则退出循环
            print(error_num,'我要退出了！')
            break
        pyautogui.click(first_loc)  # 单击
        pyautogui.scroll(-1) # 向下滚动1个单位
        continue
    
    if max_val > 0.8:  # 阈值
        print('执行了么？')
        num_loc = (max_loc[0], max_loc[1] + 130)    # 数字的坐标
        # 截取120*40的图片,先y再x
        cropped_image = screenshot[num_loc[1]:num_loc[1]+22, num_loc[0]:num_loc[0]+120]

        # 保存截图
        sc_filename = f"{list_b}.png"       # 道具名+.png
        sc_filepath = os.path.join(screenshot_folder, sc_filename)
        cv2.imwrite(sc_filepath, cropped_image)
        print(sc_filepath)
    else:
        continue    # 没扫描到，则继续下一次循环



import os
import cnocr


current_dir = os.getcwd()
print(current_dir)

# cnocr
det_model_name = "ch_PP-OCRv3_det"
det_root = os.path.join(current_dir, "model/cnocr")
# print(det_root)
# cnstd
rec_model_name = "densenet_lite_114-fc"
rec_root = os.path.join(current_dir, "model/cnstd")
# print(rec_root)

# 文字
# ocr = cnocr.CnOcr(det_model_name=det_model_name, rec_model_name=rec_model_name, det_root=det_root, rec_root=rec_root)
# 数字
# ocr = cnocr.CnOcr(det_model_name=det_model_name, rec_model_name=rec_model_name, det_root=det_root, rec_root=rec_root, cand_alphabet='0123456789')
#self.ocr = CnOcr(det_model_name='db_resnet34', rec_model_name='densenet_lite_114-fc')

# 进行文字识别
# text = ocr.ocr('image_path/123.png')

# text = ocr.ocr_for_single_line('image_path/2333.png')
# print(text)


listimg = []
# 遍历 ./screenshot 目录下的所有图片
for image_name in os.listdir(screenshot_folder):
    # 构建完整的图片路径
    screen_img_file = os.path.join(screenshot_folder, image_name)

    # 使用 OCR 模型进行文本识别
    ocr = cnocr.CnOcr(det_model_name=det_model_name, rec_model_name=rec_model_name, det_root=det_root, rec_root=rec_root, cand_alphabet='0123456789')

    '''
    这里有图像的方框坐标系，4个点
    [{'text': '1854', 'score': 0.9192507266998291, 'position': array([[         42,        
    3],
        [         90,           3],
        [         90,          19],
        [         43,          20]], dtype=float32)}]
        
    text = ocr.ocr(image_path)
    print(text)
    '''

    # 获取识别结果中的文本值
    text = ocr.ocr_for_single_line(screen_img_file)

    img_ocr_score = text['score']
    if img_ocr_score < 0.7:
        print('该文件未匹配到值',screen_img_file)
        continue
    
    # 输出图片名和文本值
    print("图片名:", image_name)
    print("文本值:", text['text'])
    print()
    image_file = image_name.rsplit(".", 1)[0]       # 读取道具名=文件名
    
    '''
    import configparser
    # 读取配置文件
    config = configparser.ConfigParser()
    config.read('config.ini', encoding='utf-8')
    '''

    name = config['lag']['name']    # 选择语言，配置块
    # 判断name是否为cn
    if config['lag']['name'] == name:
        # 读取cn配置块中的内容
        next_config = config[name]
        # 判断是否有键名与image_file相同的键
        if image_file in next_config:
            # 获取该键名的键值
            prop_name = next_config[image_file]
        else:
            prop_name = image_file      # 如果ini中未能找到，则直接输出为道具名【图片名】
    
    number = text['text']           # 读取text字典中键名为text的值
    listimg.append([prop_name, number])
    print(listimg)

#  第一种方法，csv模式
''' csv模式
import csv
filename = 'output.csv'

# w=写入，a=追加，r=只读，b=二进制，t=文本模式，rb二进制打开
with open(filename, 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerow(['道具名', '值'])  # 写入表头
    writer.writerows(listimg)  # 逐行写入数据
'''

# 第二种方法，覆盖模式填入数据
'''
import openpyxl
filename = 'output.xlsx'
# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()      # 创建模式，覆盖模式
# workbook = openpyxl.load_workbook(filename) # 读取模式，追加模式
# 获取默认的工作表
sheet = workbook.active
new_sheet_name = '养成道具'
sheet.title = new_sheet_name        # 修改工作表名
# 写入表头
sheet.append(['道具名', '值'])
# 逐行写入数据
for row in listimg:
    sheet.append(row)
# 获取B列的单元格范围
column_b = sheet['B']
# 遍历B列的单元格，并将其数据类型更改为数字类型（仅限数字单元格）
for cell in column_b:
    try:
        value = float(cell.value)
        cell.number_format = '0'  # 设置数字格式，可以根据需要进行调整
        cell.value = value
    except (ValueError, TypeError):
        pass  # 忽略非数字单元格
# 保存Excel文件
workbook。save(filename)
'''
# 第三种方法，将子列表第一个元素与A列值比对，该模式需要跟config.ini中一起改，否则不识别其他语言
import openpyxl
filename = 'output.xlsx'
# sheetname = '养成道具'
# 打开 Excel 文件
workbook = openpyxl.load_workbook(filename)
# 选择指定的工作表
sheet = workbook[sheetname]

# 遍历 A 列和 B 列的单元格
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
    cell_a = row[0]  # A 列单元格
    cell_b = row[1]  # B 列单元格
    cell_c = row[2]  # C 列单元格
    
    # 检查 A 列和 B 列单元格的值是否与 listimg 中的对应元素相同
    for row_data in listimg:
        if cell_a.value == row_data[0] or cell_b.value == row_data[0] or cell_c.value == row_data[0] :    # 该行A\B的值与
            # 获取对应行的 C 列单元格并写入子列表中的第二个元素的值
            cell_d = row[3]  # D 列单元格
            cell_d.value = row_data[1]  # 将子列表中的第二个元素的值写入 C 列单元格
            break  # 找到匹配后跳出内层循环，继续下一行的比较
            
# 获取B列的单元格范围
column_b = sheet['C']
# 遍历B列的单元格，并将其数据类型更改为数字类型（仅限数字单元格）
for cell in column_b:
    try:
        value = float(cell.value)
        cell.number_format = '0'  # 设置数字格式，可以根据需要进行调整
        cell.value = value
    except (ValueError, TypeError):
        pass  # 忽略非数字单元格
# 保存修改后的 Excel 文件
workbook。save(filename)
