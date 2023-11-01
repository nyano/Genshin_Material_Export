# 该工具用于采集图片和文字识别获取 - 背包素材、文字-数量采集
# 设为通用
# 01. 【标准部件】判断是否前置窗口为原神get_window_handle()  get_width()
# 02. 【标准部件】利用pyautogui对当前窗口（全屏）进行截取，获取窗口化的窗口的实际坐标进行计算 软件部分截图：py_screenshot(screenshot_png,pic_save=True) 全屏截图：py_all_screenshot(screenshot_png,pic_save=True)
# 03. 【标准部件】对截图的指定部分进行截取crop_and_save_image(screenshot, x, y, width, height, save_path)
# 04. 【标准部件】对屏幕指定位置进行单击和中键移动py_click(x,y,width,height) py_scroll(rows)
# 05. 【标准部件】图片与图片进行比对，确定被比对的图在截图中的位置 find_image_coordinates(image_path, template_path)
# 06. 【标准部件】文字识别 text_ocr(pic_ocr)
# 07. 测试：对单个文字进行获取
# 07. 实操：物产志中循环文字提取并输出图片

"""
BUG: 1.
待办： 
1.【完成2023-10-30】没有申请权限，没有权限没法进行滚动和点击等【2023-10-28】
2.【完成2023-10-31】滚动第2次直接滚动次数=100【2023-10-28】
    【完成2023-11-1】-测试循环完成后未执行最后循环外的内容，为什么【2023-10-31】
3. 4K屏下似乎还没有做好优化【2023-10-28】
4.【完成2023-10-30】py_click模块自带缩放，前期脚本全部又加了一遍宽高，需要取消【2023-10-28】
5.【完成2023-10-29】库进行升级备份【2023-10-28】
    - 等待库备份整理脚本
6.【完成2023-10-31】添加开始时间和结束时间，excel中添加时间【2023-10-28】
7. 支持非1080P分辨率【2023-10-28】
8.【完成2023-10-31】单击时点2下，并且等待0.5，因为图片会放大一下，不要影响截图【2023-10-28】
9.【完成2023-10-31】部分按需进行数据采集，开始、结束位置定位【2023-10-29】y轴坐标定位：utils_crop_height_width 数字定位utils_crop_white_num
10.指定部分按需进行数据采集 -【养成道具】角色经验素材、角色与武器培养素材、角色培养素材、角色突破素材、角色天赋素材、武器突破素材、【材料】素材、锻造用矿石、蒙德区域特产、……、食材、鱼饵、鱼、【2023-10-31】=》但是如果速度快的话还需要么？

新增功能：
1.添加了结束位字符【其实跟F3一致，但是自动】【2023-11-01】
2.添加循环终止符all_end 【2023-11-01】
"""

import os
import time
import shutil
import keyboard
from PIL import Image
import cnocr
import pyuac
from utils_crop_height_width import check_pixel_compliance_height
from utils_crop_height_width import check_pixel_compliance_width
from utils_crop_white_num import process_image_white_num


try:
    if not pyuac.isUserAdmin():
        pyuac.runAsAdmin()
    else:
        pass
except:
    os._exit(1)

# 输出开始时间
start_time = time.time()
formatted_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(start_time))
print("开始时间:", formatted_time)


def on_key_release(event):
    if event.name == 'f3':
        # 按下 F3 键后终止程序
        print("按下 F3 键，程序退出。")
        os._exit(0)

# 注册按键释放事件的回调函数
keyboard.on_release(on_key_release)


#'''
# print('执行01步骤***************************************')
# print('########### 01. 【标准部件】判断是否前置窗口为原神')
from utils_find_process import get_window_handle
from utils_find_process import get_window_handle_name
from utils_find_process import get_foreground_window_info
from utils_process_x_y import get_width
from utils_process_x_y import get_process_dimensions

# 其他语言的标题不是原神！【【【【【【【【【【【【备忘问题】】】】】】】】】】】】
if get_window_handle('原神'):   # 将标题名为原神的进程前置
    # print('你都没开原神！还是改语言了？')
    lag = 'chinese'
else:
    get_window_handle_name('YuanShen.exe')
    # time.sleep(0.5)
    handle, pid, name, title = get_foreground_window_info()
    if title == 'Genshin Impact':
        # print('not open Genshin Impact?')
        lag = 'english'
    else:
        lag = 'other'
    # os._exit(1)     # 没做其他语言判断
print(lag)
if lag == 'chinese':
    real_width , real_height ,scaling ,borderless = get_width('原神')
elif lag == 'english':
    real_width , real_height ,scaling ,borderless = get_process_dimensions("YuanShen.exe")
else:
    real_width , real_height ,scaling ,borderless = get_process_dimensions("YuanShen.exe")      # 其他语言
# time.sleep(0.5)
if not real_width == 1920 or not real_height == 1080:
    print('窗口不是1920X1080')
    os._exit(1)
#'''

# print('执行02步骤***************************************')
# print('########### 02. 【标准部件】利用pyautogui对当前窗口进行截取，获取窗口化的窗口的实际坐标进行计算')
# 问题：窗口必须在主界面，否则无法截取，显示为全黑
import pygetwindow as gw
import pyautogui
# 获取当前活动窗口
active_window = gw.getActiveWindow()            # <Win32Window left="0", top="0", width="1920", height="1080", title="原神">    # 偏移和尺寸
# 获取窗口位置和大小
window_left = active_window.left        # 相对于屏幕左上角的水平坐标
window_top = active_window.top          # 相对于屏幕左上角的垂直坐标
window_width = active_window.width      # 当前窗口的宽度
window_height = active_window.height    # 当前窗口的高度
# 进行截图
screenshot = pyautogui.screenshot(region=(window_left, window_top, window_width, window_height))        # 元组，window_left和window_top设为0则为全屏截图
# 当前窗口信息 <Win32Window left="1912", top="-8", width="1936", height="1056", title="dev2.py - 100 - Visual Studio Code [管理员]">
# 多屏下，为1屏时，x正确。在其他屏时，相对1屏的位置（从1屏开始算轴，不论1屏的位置）
print('当前窗口信息',active_window,'相对于屏幕左上角的水平坐标',window_left,'相对于屏幕左上角的垂直坐标',window_top,'当前窗口的宽度',window_width,'当前窗口的高度',window_height)

def py_screenshot(screenshot_png,pic_save=True):  # 截图：'test-screenshot.png'，仅限软件截图而非全屏截图,做坐标判断时，需要加入全屏偏移
    screenshot_pic = pyautogui.screenshot(region=(window_left, window_top, window_width, window_height))
    if pic_save == True:
        screenshot_pic.save(screenshot_png)     # 保存到图片，find_image_coordinates方法需要使用图片路径，否则需要加上判断
    return screenshot_pic       # 注意！输出的是图片的值，而不是图片的路径
# py_screenshot('test-screenshot.png')

def py_all_screenshot(screenshot_png,pic_save=True):  # 截图：'test-screenshot.png' ,全屏而不是单个软件部分，用于单击坐标获取
    screenshot_pic = pyautogui.screenshot()
    if pic_save == True:
        screenshot_pic.save(screenshot_png)     # 保存到图片
    return screenshot_pic       # 注意！输出的是图片的值，而不是图片的路径
# py_screenshot('test-screenshot.png')

# print('执行03步骤***************************************')
# print('########### 03. 【标准部件】对截图的指定部分进行截取')
def crop_and_save_image(screenshot, x, y, width, height, save_path):        # 从前端获得A. screenshot截图，B C.定义x、y坐标，D E.宽高矩形，F. 保存位置
    # 判断输入类型是图像路径还是图像对象
    if isinstance(screenshot, str):  # 如果是字符串，则加载图像
        screenshot = Image.open(screenshot)
    # 根据给定的坐标和尺寸进行截取
    cropped_image = screenshot.crop((x, y, x + width, y + height))
    # 保存截取的图片
    cropped_image.save(save_path)
# crop_and_save_image(screenshot, 140, 125, 114, 114, 'test-cropped_image.png')

# print('执行04步骤***************************************')
# print('########### 04. 【标准部件】对屏幕指定位置进行单击和中键移动')
# 单击操作
def py_click(x,y,width,height):     # x坐标，y坐标=左上角顶点坐标，width、height=宽高中心点增值，也可以设为0
    click_loc = (x + width + window_left,y + height + window_top)    # 左上角顶点+中心点增值 + 窗口与屏幕的差距
    # print(x,y)
    # print(width,height)
    # print(window_left,window_top)
    # print(click_loc)
    # print('单击位置',click_loc)
    pyautogui.click(click_loc) 
    pyautogui.click(click_loc)  # 单击两下，以便去掉【新】字
    # time.sleep(0.5)             # 等待一下，以便让图片能够动画完成
    return click_loc
# py_click(140, 125, 114/2, 114/2)

# 滚动操作
def py_scroll(rows):
    for i in range(0,rows):                    # 右侧为绕过的行数（含本行）,物产志提瓦特物产首页5行-堇瓜
        for i in range(0,9):                # 素材的一行，物产志和素材一致,9次滚动
            pyautogui.scroll(-1,x=0,y=0)    # 只定义正值（向上）和负值（向下），x水平正值向右，负值向左。y正值向上，负值向下
    pyautogui.scroll(1,x=0,y=0)             # 返回向上一次
    # time.sleep(0.5)
# py_scroll(5)    

# print('执行05步骤***************************************')
# print('########### 05. 【标准部件】图片与图片进行比对，确定被比对的图在截图中的位置')
import cv2
import numpy as np
def find_image_coordinates(image_path, template_path):      # image_path = 屏幕截图 template_path = 对比图,添加了，输入内容判断是图片还是图片的路径
    # 判断image_path和template_path是否为字符串路径
    if isinstance(image_path, str):
        # 读取图像文件
        image = cv2.imread(image_path)
    else:
        # image_path为图像变量
        image = np.array(image_path)
        image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)

    if isinstance(template_path, str):
        # 读取模板文件
        template = cv2.imread(template_path)
    else:
        # template_path为图像变量
        template = np.array(template_path)
        template = cv2.cvtColor(template, cv2.COLOR_RGB2BGR)
    # 使用模板匹配算法查找模板在图像中的位置
    result = cv2.matchTemplate(image, template, cv2.TM_CCOEFF_NORMED)
    _, max_val, _, max_loc = cv2.minMaxLoc(result)      # max_val是相似度，max_loc为坐标（元组）
    # 获取模板的宽度和高度
    template_width = template.shape[1]
    template_height = template.shape[0]

    # 计算模板的左上角和右下角坐标
    top_left = max_loc      # 左上角坐标
    bottom_right = (top_left[0] + template_width, top_left[1] + template_height)    # 右下角坐标，依赖比对图的大小，如果有差异、拉伸则错误，114或125
    print('比对相似度为：',max_val, '坐标为：',top_left, bottom_right)
    return max_val,top_left, bottom_right

# print('执行06步骤***************************************')
# print('########### 06. 【标准部件】文字识别')
def text_ocr(pic_ocr,single=False,num =False):      # 是否为单行，是否为纯数字 ，一定要注意单行和多行输出的结果不同！
    """未测试，似乎必须为路径，不知道图像对象行不行。路径和文件名必须为全英文，但是相对路径不要紧"""
    current_dir = os.getcwd()   # 当前目录，要把模型放在当前目录，model目录
    if lag == 'english':
        det_model_name = "en_PP-OCRv3_det"
    elif lag == 'chinese':
        det_model_name = "ch_PP-OCRv3_det"
    else:
        det_model_name = "en_PP-OCRv3_det"              # 【【【【【注意！其他语言没有指定OCR模型！可能全错】】】】】
    det_root = os.path.join(current_dir, "model/cnocr")
    rec_model_name = "densenet_lite_114-fc"
    rec_root = os.path.join(current_dir, "model/cnstd")
    # 使用 OCR 模型进行文本识别，仅数字后面加上, cand_alphabet='0123456789'
    if num == False:
        ocr = cnocr.CnOcr(det_model_name=det_model_name, rec_model_name=rec_model_name, det_root=det_root, rec_root=rec_root)
    else:
        ocr = cnocr.CnOcr(det_model_name=det_model_name, rec_model_name=rec_model_name, det_root=det_root, rec_root=rec_root,cand_alphabet='0123456789')
    if single == True:
        # 获取识别结果中的文本值 # 单行内容  
        text = ocr.ocr_for_single_line(pic_ocr)         # 单行ocr_word = textocr['text']
    else:
        text = ocr.ocr(pic_ocr)                         # 多行ocr_word = textocr[0]['text']  数组 {'text': '养成道具', 'score': 0.9842932820320129}
    return text

"""【标准部件】文字识别-执行：截图、截取文字部分、识别输出
screenshot_pic = py_screenshot(f'./{folder_name}/{screenshot_png}')                                 # 截图
crop_and_save_image(screenshot_pic, 130, 30, 250, 40, f'./{folder_name}/{crop_cnocr_png}')          # 截取文字部分图片
textocr = text_ocr(f'./{folder_name}/{crop_cnocr_png}',single=True)                                 # 自动识别了中英文
type_list = textocr['text']         # 单行（养成道具、材料）
"""
"""【标准部件】图像识别定位
screenshot_pic = py_screenshot(f'./{folder_name}/{screenshot_png}.png')
val, left_top_scoll, _ = find_image_coordinates(f'./{folder_name}/{screenshot_png}.png',f'./{folder_name}/{crop_cnocr_png}-scoll.png') 
if val > 0.8:
    left_top_y = left_top_scoll[1]  # 确认到Y轴坐标
"""


# 变量整合
# window_left 鼠标定位时窗口x轴，window_top鼠标定位时窗口y轴
folder_name = "dev4"
screenshot_png = 'A-001-test-screenshot'
crop_cnocr_png = 'A-002-test-crop-cnocr'
type_list = '养成道具'
left_top_x = 118                # 第一排第一个x轴
left_top_y = 121                # 第一排第一个y轴，竖向移动175
column_x = 0                    # 横向移动+147
right_word_name = ''            # 右1-文字-名称
right_word_name_score = ''      # 右1-文字-名称的ocr分数
right_word_type = ''            # 右2-文字-类别
right_word_type_score = ''      # 右2-文字-类别的ocr分数
word_num = ''                   # 数字
word_num_score = ''             # 数字的ocr分数
nums = 0                        # 总数量
rows = 0        # 行数
cols = 0        # 列数
scroll_error = 0    # 滚动后发生错误
last_word = ''          # 上一个文字，右1-名称，最后一排比对用
scoll_num = 0           # 滚动后次数，如果大于2，相同，则代表为最后一行了
scoll = 99              # 设定的滚动次数
last_row = False        # 标记为最后一行
first_scoll_y = 0       # 最后一次滚动循环的第一次滚动y轴
cut_word = True         # 结束点标记，最后一个就终止
all_end = False         # 真-最后结束点

if lag == 'english':        # excel 列定位
    ex_type_list = 'A'
    ex_a = 'C'              # 名称【右1-文字】right_word_name
    ex_b = 'D'              # 类目【右2-类目】right_word_type
    ex_c = 'E'              # 数字【数字】word_num
elif lag == 'chinese':
    ex_type_list = 'B'
    ex_a = 'F'
    ex_b = 'G'
    ex_c = 'H'
else:
    # lag = 'other'
    ex_a = 'I'
    ex_b = 'J'
    ex_c = 'K'

from openpyxl import Workbook, load_workbook
import os
if os.path.exists(folder_name):
    shutil.rmtree(folder_name)      # 清空和重建文件夹
os.mkdir(folder_name)

examples_folder_name = f"./{folder_name}/examples"      # ocr学习目录
if os.path.exists(examples_folder_name):
    shutil.rmtree(folder_name)      # 清空和重建文件夹
os.mkdir(examples_folder_name)

filename = f'./{folder_name}/test.xlsx'
if not os.path.isfile(filename):
    workbook = Workbook()   # 创建一个新的工作簿
    worksheet = workbook.active # 选择第一个工作表
    workbook.save(filename) # 保存新建的 Excel 文件
# print('执行07步骤***************************************')
'''背包 -养成道具、材料
A标题文字定位：养成道具、材料：(130,30) 250x40
B一排1：(118,121)  123x123
C一排2：(265,121)   左右偏移【147】
D二排1：(118,296)   上下偏移【175】
一排8个 +147  最后一个1147
一个界面4+1（第5显示一半）
E文字-右1【名称】-大概率单行：【多行：奇械发条备件·科培琉司】 (1315,127) 482x50 (字可能会侵入下方)
F文字-右2【分类】-单行：(1315,180) 265x40
G数字-【数量】-单行：(118,244)124x25  Y轴+123
'''
"""
1.A【单行】对标题(130,30) 250x40部分进行判断，确定类目
2.E【单行】名称（右1文字识别）(1334,127)  (部分有多行)、截图（440x45）用于OCR学习 命名：A-总数-排数-列数-名称-分类.png【英文有2行，y上+5，下+6=》未加入】
3.F【单行】分类（右2文字识别）(1334,180) 、截图（245x40）用于OCR学习 ，只有4类？ 命名：B-总数-排数-列数-名称-分类.png【英文有2行，要y+15=》未加入】
4.G【单行】数量（数字识别） x+10,Y轴+128 ，(128,249)截图（100x25），用于OCR学习，该值由于Y轴坐标差异，有偏移
5.B 首位定位(118,121)，A.截图(123x123) 命名：总数-排数-列数-名称-分类.png
6.加入到excel表格中
循环2-5，一行8个，3行，+1行，滚动3行，重新定位
990.X轴循环，7次（8个），每次+147 column_x
991.Y轴循环，3排+第一个截图
992.滚动，循环，滚动，重新截图-定位
993.最后一排判断
"""

"""992.滚动，循环，滚动，重新截图-定位"""
for sc in range(1, scoll + 1):  # 滚动
    if all_end == True:
        break
    if sc >1:   # 进行Y轴定位
        screenshot_pic = py_screenshot(f'./{folder_name}/{screenshot_png}.png')
        val, left_top_scoll, _ = find_image_coordinates(f'./{folder_name}/{screenshot_png}.png',f'./{folder_name}/{crop_cnocr_png}-scoll.png') 
        if val > 0.8:
            left_top_y = left_top_scoll[1]  # 确认到Y轴坐标
        """该循环表示如果y轴大于175，则进行重复滚动，保持滚动对象在最上方，上方极限是125（会看不到），下方极限是330（扫不到第四行）"""
        while left_top_y > 175:
            for i in range(0,2):                # 素材的一行，物产志和素材一致,9次滚动
                pyautogui.scroll(-1,x=0,y=0)    # 只定义正值（向上）和负值（向下），x水平正值向右，负值向左。y正值向上，负值向下
            # time.sleep(0.5)
            screenshot_pic = py_screenshot(f'./{folder_name}/{screenshot_png}.png')
            val, left_top_scoll, _ = find_image_coordinates(f'./{folder_name}/{screenshot_png}.png',f'./{folder_name}/{crop_cnocr_png}-scoll.png') 
            if val > 0.8:
                left_top_y = left_top_scoll[1]  # 确认到Y轴坐标
                scoll_num += 1
                if left_top_y >= 365 or abs(first_scoll_y - left_top_scoll[1]) < 10:        # 高度大于365或者2次滚动后的y轴对比，小于10，无法滚动后一般应该相等
                    scoll_num += 1
                    if scoll_num >= 2:
                        last_row = True
                        break           # 属于最后一行了，退出循环while，需要继续，做最后一部分的循环【实际结束是文字一致时】
                else:
                    first_scoll_y = left_top_scoll[1]

            else:   # 没找到，可能过了，那么向上移动
                for i in range(0,5):                # 素材的一行，物产志和素材一致,9次滚动
                    pyautogui.scroll(1,x=0,y=0)    # 只定义正值（向上）和负值（向下），x水平正值向右，负值向左。y正值向上，负值向下
                    scroll_error += 1
                    if scroll_error == 3:
                        print('滚动错误，未找到')
                        os._exit(1)
                
    """991.Y轴循环，3排+第一个截图"""
    for row in range(1, 3 + 1): # 3行
        if all_end == True:
            break
        rows += 1       # 行数
        cols = 0        # 清空列数
        column_x = 0    # 重置x轴增值
        

        """990.横向循环，7次（8个），每次+147 column_x"""
        for col_list in range(1, 8 + 1):    # 8次
            if all_end == True:
                break
            nums += 1
            cols += 1
            left_top_x = 118 + column_x
            
            """1.A【单行】对标题(130,30) 250x40部分进行判断，确定类目"""
            """
            screenshot_pic = py_screenshot(f'./{folder_name}/{screenshot_png}.png')                                 # 截图
            crop_and_save_image(screenshot_pic, 130, 30, 250, 40, f'./{folder_name}/{crop_cnocr_png}-a.png')          # 截取文字部分图片
            textocr = text_ocr(f'./{folder_name}/{crop_cnocr_png}-a.png',single=True)                                 # 自动识别了中英文
            type_list = textocr['text']         # 单行（养成道具、材料）
            """

            click_loc = py_click(left_top_x , left_top_y  , 123/2, 123/2)                                         # 点击图片中间进行定位
            # print('单击位置',left_top_x , left_top_y )
            screenshot_pic = py_screenshot(f'./{folder_name}/{screenshot_png}.png')                                 # 截图

            """2.E【单行】名称（右1文字识别）(1334,127)  (部分有多行)、截图（440x45）用于OCR学习 命名：A-总数-排数-列数-名称-分类.png【英文有2行，y上+5，下+6=》未加入】"""
            crop_and_save_image(screenshot_pic, 1334, 127, 440, 45, f'./{folder_name}/{crop_cnocr_png}-b.png')          # 截取文字部分图片
            textocr = text_ocr(f'./{folder_name}/{crop_cnocr_png}-b.png',single=True)                                   # 识别右1文字-名称
            right_word_name = textocr['text']
            right_word_name_score = round(textocr['score'], 2)      # 只有0.57
            # print(right_word_name, right_word_name_score)


            """3.F【单行】分类（右2文字识别）(1334,180) 、截图（245x40）用于OCR学习 ，只有4类？ 命名：B-总数-排数-列数-名称-分类.png【英文有2行，要y+15=》未加入】"""
            crop_and_save_image(screenshot_pic, 1334, 180, 245, 40, f'./{folder_name}/{crop_cnocr_png}-c.png')          # 截取文字部分图片
            textocr = text_ocr(f'./{folder_name}/{crop_cnocr_png}-c.png',single=True)                                   # 识别右1文字-名称
            right_word_type = textocr['text']
            right_word_type_score = round(textocr['score'], 2)      # 只有0.57
            # print(right_word_type, right_word_type_score)


            """4.G【单行】数量（数字识别） x+10,Y轴+128 ，(128,249)截图（100x25），用于OCR学习，该值由于Y轴坐标差异，有偏移
            字的白色部分色系(233,229,220) 字的白色系部分高度27，满色24px x轴 122px"""
            crop_and_save_image(screenshot_pic, left_top_x, left_top_y + 120, 123, 50, f'./{folder_name}/{crop_cnocr_png}-d.png')          # 截取文字部分图片
            process_image_white_num(f'./{folder_name}/{crop_cnocr_png}-d.png',f'./{folder_name}/{crop_cnocr_png}-d.png')
            textocr = text_ocr(f'./{folder_name}/{crop_cnocr_png}-d.png',single=True,num=True)                                   # 识别右1文字-名称
            word_num = textocr['text']
            word_num_score = round(textocr['score'], 2)      # 0.95
            # print(word_num, word_num_score)
            # print('核实数字的坐标',(left_top_x + 10, left_top_y + 128))

            print(f'总数：{nums},行：{rows},列:{cols},名：{right_word_name},分类:{right_word_type},值：{word_num},坐标:({left_top_x},{left_top_y}),滚动次数:{sc -1}')   # 滚动-1使得0开始

            
            """6.加入到excel表格中"""
            workbook = load_workbook(filename)
            worksheet = workbook.active                 # 选择第一个工作表
            column = worksheet[ex_type_list]                           # 查找第一列中的空白行
            empty_row = len(column) + 1                         # 追加到第一列的空白行
            # worksheet[f'{ex_type_list}{empty_row}'] = type_list
            worksheet[f'{ex_a}{empty_row}'] = right_word_name           # 第一列,【名称】
            worksheet[f'{ex_b}{empty_row}'] = right_word_type           # 第二列，【类目】
            worksheet[f'{ex_c}{empty_row}'] = word_num                  # 第三类，【数字】
            workbook.save(filename)
            
            """最后一行比对，针对文字进行判断"""
            if cut_word == True:
                cut_last_word = '子探测单元'
                last_word = cut_last_word
                # right_word_name = '子探测单元'
                last_row = True
            
            if last_word == right_word_name:
                print('已到最后一行，或比对错误，退出')
                if last_row == True:
                    print('核实确实为最后一行，退出')
                    sc = scoll + 2
                    # 输出结束时间
                    end_time = time.time()
                    end_formatted_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(end_time))
                    print("结束时间:", end_formatted_time)
                    # 计算并输出总运行时间
                    total_time = end_time - start_time
                    minutes = int(total_time // 60)
                    seconds = int(total_time % 60)
                    run_time = str(minutes) + "分" + str(seconds) + "秒"
                    print(f"总运行时间: {run_time}")
                    
                    # 在excel中添加运行时间
                    workbook = load_workbook(filename)
                    worksheet = workbook.active                 # 选择第一个工作表
                    column = worksheet[ex_type_list]                           # 查找第一列中的空白行
                    empty_row = len(column) + 1                         # 追加到第一列的空白行
                    # worksheet[f'{ex_type_list}{empty_row}'] = type_list
                    worksheet[f'{ex_a}{empty_row}'] = formatted_time           # 第一列,【名称】 -开始时间
                    worksheet[f'{ex_b}{empty_row}'] = end_formatted_time           # 第二列，【类目】 - 结束时间
                    worksheet[f'{ex_c}{empty_row}'] = run_time                  # 第三类，【数字】 -运行时间
                    workbook.save(filename)
                    
                    all_end = True
                    break
                else:
                    print('对比错误？可能有相同值，退出')
                # os._exit(0)     # 或者改为退出循环
            else:
                last_word = right_word_name

            # 上面3个图全部复制一份用于OCR学习
            # shutil.copy2(f'./{folder_name}/{crop_cnocr_png}-b.png',f'./{folder_name}/examples/B-{nums}-{rows}-{cols}-{right_word_type}-{right_word_name}.png')
            # shutil.copy2(f'./{folder_name}/{crop_cnocr_png}-c.png',f'./{folder_name}/examples/C-{nums}-{rows}-{cols}-{right_word_name}-{right_word_type}.png')
            # shutil.copy2(f'./{folder_name}/{crop_cnocr_png}-d.png',f'./{folder_name}/examples/D-{nums}-{word_num}.png')

            """5.B 首位定位(118,121)，A.截图(123x123) 命名：总数-排数-列数-分类-名称.png"""

            # crop_and_save_image(screenshot_pic, left_top_x, left_top_y, 123, 123, f'./{folder_name}/{nums}-{rows}-{cols}-{right_word_type}-{right_word_name}.png')       # 截取图片保存（但是好像没啥用了）


            
            column_x += 147  # 每次x轴移动147
        left_top_y += 175
        
    """滚动前定位，截取下一排的第一个图片，用于定位，X=118，Y+= 175,123x65(Y的一半左右)"""
    click_loc = py_click(118, left_top_y , 123/2, 123/2)
    print('单击位置',118 , left_top_y)
    crop_and_save_image(screenshot_pic, 118, left_top_y , 123, 60, f'./{folder_name}/{crop_cnocr_png}-scoll.png')
    check_pixel_compliance_height(f'./{folder_name}/{crop_cnocr_png}-scoll.png', 0, 0)      # 裁掉上部分，y轴
    check_pixel_compliance_width(f'./{folder_name}/{crop_cnocr_png}-scoll.png', 0, 0)       # 裁掉左部分，但是对于滚动好像不需要，x轴，因为固定
    print('开始滚动')
    py_scroll(3)    # 滚动，3行，30次

print('我执行了么？')




"""多线程
1.首先遍历背包，点击
2.截屏，截图a.文字部分 b.数字部分，将路径保存为数组列表[{编号,文字部分,数字部分}]
3.循环X轴移动8次 ，单次循环完成后，将列表发给子进程进行图像识别
4.循环Y轴移动3次
5.Y轴循环，3排+第一个截图
992.滚动，循环，滚动，重新截图-定位
993.最后一排判断

第二套
1.遍历背包、点击、滚动、截图，将所有图片保存，图片名：001-顶点坐标.png
2.然后获得文件夹内所有图片的路径，8个一组
3.多线程进行OCR
"""


"""
BUG :1.Y轴移动有一定距离的偏移，无法准确定位
"""

